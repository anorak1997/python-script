import openpyxl
import email
import imaplib
import getpass
from datetime import datetime
import ezodf


def check_and_save_job_emails(email_address, password):
    """Checks for new job emails and saves them to an ODS file."""

    # Connect to the IMAP server
    mail = imaplib.IMAP4_SSL('imap.gmail.com')

    try:
        # Log in to your email account
        mail.login(email_address, password)

        # Select the mailbox you want to check (e.g., 'inbox')
        mail.select('inbox')

        # Search for all emails with the 'job' keyword in the subject and received today
        today = datetime.today().strftime('%d-%b-%Y')
        search_criteria = f'(SINCE "{today}") (SUBJECT "job")'
        status, messages = mail.search(None, search_criteria)

        # Get the list of email IDs
        email_ids = messages[0].split()

        if not email_ids:
            print("No job-related emails found today.")
            return

        print(f"Found {len(email_ids)} job-related emails today.")

        # Create a new workbook
        workbook = openpyxl.Workbook()

        # Create a new worksheet
        worksheet = workbook.create_sheet("Job Emails")

        # Set column headers
        worksheet.cell(row=1, column=1).value = 'Subject'
        worksheet.cell(row=1, column=2).value = 'From'
        worksheet.cell(row=1, column=3).value = 'Body'

        # Iterate through each email and save information to the ODS file
        row_index = 2
        for email_id in email_ids:
            status, msg_data = mail.fetch(email_id, '(RFC822)')

            # Parse the email content
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            # Get subject, sender, and body
            subject = msg['Subject']
            sender = msg['From']
            body = msg.get_payload()

            # Extract the text content of the email message
            extracted_body_text = ""
            if isinstance(body, str):  # Check if it's a plain text message
                extracted_body_text = body
            elif isinstance(body, list):  # Check if it's a list of MIME parts
                for part in body:
                    if isinstance(part, email.message.Message):
                        extracted_body_text += part.get_payload()

            # Save the extracted information to the worksheet
            worksheet.cell(row=row_index, column=1).value = subject
            worksheet.cell(row=row_index, column=2).value = sender
            worksheet.cell(row=row_index, column=3).value = extracted_body_text

            row_index += 1

        # Save the workbook
        workbook.save('job_emails.ods')

        print("Job-related emails saved to job_emails.ods")

    finally:
        # Logout and close the connection
        mail.logout()

if __name__ == "__main__":
    # Input your email address and password
    email_address = "siddharthadubey.1997@gmail.com"
    password = "wiqg hfdp vjgm bhul"

    check_and_save_job_emails(email_address, password)

