import openpyxl
import email
import imaplib
from datetime import datetime

def get_text_payload(part):
    """
    Recursively extract the text payload from MIME parts.
    """
    if isinstance(part, str):  # Check if it's a string
        return part
    elif isinstance(part, list):  # Check if it's a list of MIME parts
        return ''.join(get_text_payload(subpart) for subpart in part if subpart)
    elif isinstance(part, email.message.Message):  # Check if it's a MIME part
        return get_text_payload(part.get_payload())
    else:
        return ""

def check_and_save_job_emails(email_address, password):
    """Checks for all emails received today and saves them to an ODS file."""

    # Connect to the IMAP server
    mail = imaplib.IMAP4_SSL('imap.gmail.com')

    try:
        # Log in to your email account
        mail.login(email_address, password)

        # Select the mailbox you want to check (e.g., 'inbox')
        mail.select('inbox')

        # Search for all emails received today
        today = datetime.today().strftime('%d-%b-%Y')
        search_criteria = f'(SINCE "{today}")'
        status, messages = mail.search(None, search_criteria)

        # Get the list of email IDs
        email_ids = messages[0].split()

        if not email_ids:
            print("No emails found today.")
            return

        print(f"Found {len(email_ids)} emails today.")

        # Create a new workbook
        workbook = openpyxl.Workbook()

        # Create a new worksheet
        worksheet = workbook.create_sheet("All Emails")

        # Set column headers
        worksheet.cell(row=1, column=1).value = 'Subject'
        worksheet.cell(row=1, column=2).value = 'From'
        worksheet.cell(row=1, column=3).value = 'Body'

        # Iterate through each email and save information to the ODS file
        row_index = 2
        for email_id in email_ids:
            status, msg_data = mail.fetch(email_id, '(RFC822)')

            # Parse the email content
            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)

            # Get subject, sender, and body
            subject = msg['Subject']
            sender = msg['From']
            body = msg.get_payload()

            # Extract the text content of the email message
            extracted_body_text = get_text_payload(body)

            # Save the extracted information to the worksheet
            worksheet.cell(row=row_index, column=1).value = subject
            worksheet.cell(row=row_index, column=2).value = sender
            worksheet.cell(row=row_index, column=3).value = extracted_body_text

            row_index += 1

        # Save the workbook
        workbook.save('all_emails.ods')

        print("All emails received today saved to all_emails.ods")

    except imaplib.IMAP4.error as e:
        print(f"Failed to log in. Check your credentials. Error: {e}")

    finally:
        # Logout and close the connection
        mail.logout()

if __name__ == "__main__":
    # Input your email address and password
    email_address = "siddharthadubey.1997@gmail.com"
    password = "wiqg hfdp vjgm bhul"

    check_and_save_job_emails(email_address, password)

    

